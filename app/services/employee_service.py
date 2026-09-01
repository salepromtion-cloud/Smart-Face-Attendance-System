"""
Employee Service
================

Read-only access to the Employee_Master sheet inside the shared
attendance workbook, plus permanent Employee_ID generation.

Scope of this module:
    - Read existing employee records from Employee_Master.
    - Generate the next permanent, sequential Employee_ID.
    - Look employees up by name or by Employee_ID.

Out of scope (handled elsewhere in the project):
    - Writing to Employee_Master.
    - Creating or repairing the workbook.
    - Registration workflow, attendance marking, face recognition,
      Flask routes, or any HTML.

Employee_ID format: MP0001, MP0002, MP0003, ...
IDs are permanent. Once assigned, an ID is never reused, even if the
corresponding employee later leaves or is deactivated. To guarantee
this, get_next_employee_id() always inspects every Employee_ID
currently present in Employee_Master (active and inactive alike) and
returns (highest valid number found) + 1. Gaps are never backfilled.

Date format used across the project: DD-MM-YYYY
(Registered_Date is written by the registration workflow, not here.)
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

# This file lives at: app/services/employee_service.py
# parents[0] -> app/services
# parents[1] -> app
# parents[2] -> project root
PROJECT_ROOT: Path = Path(__file__).resolve().parents[2]
EXCEL_PATH: Path = PROJECT_ROOT / "data" / "attendance_data.xlsx"

EMPLOYEE_SHEET_NAME = "Employee_Master"
EMPLOYEE_ID_PREFIX = "MP"
EMPLOYEE_ID_DIGITS = 4  # MP + 4 digits -> MP0001


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class EmployeeServiceError(Exception):
    """Base exception for all employee_service failures."""


class WorkbookNotFoundError(EmployeeServiceError):
    """Raised when data/attendance_data.xlsx does not exist on disk."""


class WorksheetNotFoundError(EmployeeServiceError):
    """Raised when the Employee_Master worksheet is missing from the workbook."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _open_employee_sheet():
    """
    Open attendance_data.xlsx in read-only mode and return both the
    workbook and the Employee_Master worksheet.

    The caller owns the returned workbook and MUST close it
    (e.g. via a try/finally block) once done reading.

    Raises:
        WorkbookNotFoundError: if the Excel file does not exist.
            This service never creates the workbook automatically.
        WorksheetNotFoundError: if the Employee_Master sheet is missing.
    """
    if not EXCEL_PATH.exists():
        raise WorkbookNotFoundError(
            f"Attendance workbook not found at: {EXCEL_PATH}. "
            "This service does not create the workbook automatically — "
            "please ensure the file exists."
        )

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if EMPLOYEE_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise WorksheetNotFoundError(
            f"Worksheet '{EMPLOYEE_SHEET_NAME}' not found in {EXCEL_PATH}. "
            f"Available sheets: {workbook.sheetnames}"
        )

    return workbook, workbook[EMPLOYEE_SHEET_NAME]


def _normalize(value: Any) -> str:
    """Convert a cell value to a trimmed string; None becomes an empty string."""
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_employee_master() -> List[Dict[str, Any]]:
    """
    Read all employee records from Employee_Master.

    The first row is treated as the header row. Column names are kept
    exactly as they appear in the sheet (Employee_ID, Employee_Name,
    Department, Face_Registered, Status, Registered_Date). Any row
    that is completely empty is skipped.

    Returns:
        A list of dictionaries, one per employee row, mapping
        column name -> cell value.

    Raises:
        WorkbookNotFoundError: if attendance_data.xlsx does not exist.
        WorksheetNotFoundError: if Employee_Master sheet is missing.
    """
    workbook, worksheet = _open_employee_sheet()
    try:
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            return []

        headers = [_normalize(h) for h in header_row]

        employees: List[Dict[str, Any]] = []
        for row in rows:
            if row is None or all(cell is None for cell in row):
                continue
            employees.append(dict(zip(headers, row)))

        return employees
    finally:
        workbook.close()


def get_next_employee_id() -> str:
    """
    Determine the next permanent, sequential Employee_ID.

    Inspects the Employee_ID column of every record in Employee_Master
    (regardless of Status), safely ignores any value that is not a
    valid MP#### ID, and returns the next number after the highest
    valid ID found. IDs are never reused, so gaps are never backfilled.
    The ID is never derived from name, department, row position, or date.

    Examples:
        no employees at all         -> MP0001
        MP0001                      -> MP0002
        MP0001, MP0002, MP0005      -> MP0006
        highest valid ID is MP0081  -> MP0082

    Returns:
        The next Employee_ID, e.g. "MP0082".

    Raises:
        WorkbookNotFoundError: if attendance_data.xlsx does not exist.
        WorksheetNotFoundError: if Employee_Master sheet is missing.
    """
    employees = get_employee_master()

    highest = 0
    for employee in employees:
        raw_id = _normalize(employee.get("Employee_ID")).upper()

        if not raw_id.startswith(EMPLOYEE_ID_PREFIX):
            continue  # ignore invalid/non-MP IDs safely

        suffix = raw_id[len(EMPLOYEE_ID_PREFIX):]
        if not suffix.isdigit():
            continue  # ignore malformed IDs safely (e.g. "MPX", "MP")

        highest = max(highest, int(suffix))

    next_number = highest + 1
    return f"{EMPLOYEE_ID_PREFIX}{next_number:0{EMPLOYEE_ID_DIGITS}d}"


def get_employee_by_name(employee_name: str) -> Optional[Dict[str, Any]]:
    """
    Find an employee record by Employee_Name.

    Comparison is case-insensitive and ignores leading/trailing
    whitespace on both the input and the stored value. If the same
    name appears more than once in Employee_Master, the first match
    is returned — duplicate-name handling is left to the registration
    workflow, not decided here.

    Args:
        employee_name: The name to search for.

    Returns:
        The matching employee dictionary, or None if not found.

    Raises:
        WorkbookNotFoundError: if attendance_data.xlsx does not exist.
        WorksheetNotFoundError: if Employee_Master sheet is missing.
    """
    if not employee_name or not str(employee_name).strip():
        return None

    target = str(employee_name).strip().casefold()

    for employee in get_employee_master():
        name = _normalize(employee.get("Employee_Name")).casefold()
        if name == target:
            return employee

    return None


def get_employee_by_id(employee_id: str) -> Optional[Dict[str, Any]]:
    """
    Find an employee record by exact Employee_ID.

    Comparison normalizes leading/trailing whitespace and is
    case-insensitive (e.g. "mp0001" matches "MP0001").

    Args:
        employee_id: The Employee_ID to search for.

    Returns:
        The matching employee dictionary, or None if not found.

    Raises:
        WorkbookNotFoundError: if attendance_data.xlsx does not exist.
        WorksheetNotFoundError: if Employee_Master sheet is missing.
    """
    if not employee_id or not str(employee_id).strip():
        return None

    target = str(employee_id).strip().casefold()

    for employee in get_employee_master():
        current_id = _normalize(employee.get("Employee_ID")).casefold()
        if current_id == target:
            return employee

    return None