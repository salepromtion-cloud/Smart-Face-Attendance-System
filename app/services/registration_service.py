"""
Employee Registration Service
==============================

Writes registration data for an already-existing employee record in
Employee_Master. This module assigns a permanent Employee_ID to an
employee whose name is already present in the workbook, and marks
that employee as registered (pending face capture).

This module does NOT handle:
    - Camera access
    - Face recognition / face encoding
    - Attendance marking
    - Flask routes
    - HTML

Workbook:  data/attendance_data.xlsx
Worksheet: Employee_Master

Date format used across the project: DD-MM-YYYY
"""

from datetime import datetime
from typing import Any, Dict

from openpyxl import load_workbook

from app.services.employee_service import (
    EXCEL_PATH,
    EMPLOYEE_SHEET_NAME,
    get_employee_by_name,
    get_next_employee_id,
    EmployeeServiceError,
)

REQUIRED_COLUMNS = (
    "Employee_ID",
    "Employee_Name",
    "Department",
    "Face_Registered",
    "Status",
    "Registered_Date",
)


class RegistrationError(EmployeeServiceError):
    """Base exception for all employee-registration failures."""


def _normalize(value: Any) -> str:
    """Convert a cell value to a trimmed string; None becomes ''."""
    if value is None:
        return ""
    return str(value).strip()


def register_employee(employee_name: str) -> Dict[str, Any]:
    """
    Register an existing Employee_Master employee by assigning them a
    permanent Employee_ID.

    The employee must already exist in Employee_Master. This function
    never creates a new employee row and never overwrites an
    Employee_ID that has already been assigned.

    Args:
        employee_name: Name of an existing Employee_Master employee.
            Matching ignores leading/trailing spaces and is
            case-insensitive.

    Returns:
        A dictionary describing the newly registered employee:
        Employee_ID, Employee_Name, Department, Face_Registered,
        Status, Registered_Date.

    Raises:
        RegistrationError: for any validation, lookup, or write
            failure. Raw exceptions (file I/O, missing sheet, missing
            columns, etc.) are always wrapped in a RegistrationError
            with a clear message — never propagated directly.
    """
    name = _normalize(employee_name)
    if not name:
        raise RegistrationError("Employee name is required.")

    try:
        employee = get_employee_by_name(name)
    except EmployeeServiceError as exc:
        raise RegistrationError(f"Could not read Employee_Master: {exc}") from exc

    if employee is None:
        raise RegistrationError(f"Employee '{name}' was not found in Employee_Master.")

    existing_id = _normalize(employee.get("Employee_ID"))
    if existing_id:
        raise RegistrationError(
            f"Employee '{name}' already has Employee_ID {existing_id}."
        )

    try:
        new_employee_id = get_next_employee_id()
    except EmployeeServiceError as exc:
        raise RegistrationError(f"Could not generate Employee_ID: {exc}") from exc

    if not EXCEL_PATH.exists():
        raise RegistrationError(f"Attendance workbook not found at: {EXCEL_PATH}")

    workbook = None
    registered_date = ""
    try:
        try:
            workbook = load_workbook(EXCEL_PATH)
        except Exception as exc:
            raise RegistrationError(f"Could not open attendance workbook: {exc}") from exc

        if EMPLOYEE_SHEET_NAME not in workbook.sheetnames:
            raise RegistrationError(
                f"Worksheet '{EMPLOYEE_SHEET_NAME}' not found in {EXCEL_PATH}."
            )
        worksheet = workbook[EMPLOYEE_SHEET_NAME]

        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            raise RegistrationError(f"'{EMPLOYEE_SHEET_NAME}' has no header row.")

        columns = {
            _normalize(cell.value): cell.column
            for cell in header_row
            if _normalize(cell.value)
        }

        missing_columns = [c for c in REQUIRED_COLUMNS if c not in columns]
        if missing_columns:
            raise RegistrationError(
                f"Employee_Master is missing required column(s): {missing_columns}"
            )

        # Locate the exact employee row by Employee_Name.
        target_row = None
        for row in worksheet.iter_rows(min_row=2):
            name_cell = row[columns["Employee_Name"] - 1]
            if _normalize(name_cell.value).casefold() == name.casefold():
                target_row = name_cell.row
                break

        if target_row is None:
            raise RegistrationError(
                f"Employee '{name}' could not be located in the workbook."
            )

        # Re-check immediately before writing (concurrency safety).
        id_cell = worksheet.cell(row=target_row, column=columns["Employee_ID"])
        current_id = _normalize(id_cell.value)
        if current_id:
            raise RegistrationError(
                f"Employee '{name}' already has Employee_ID {current_id}."
            )

        registered_date = datetime.now().strftime("%d-%m-%Y")

        id_cell.value = new_employee_id
        worksheet.cell(row=target_row, column=columns["Face_Registered"]).value = "No"
        worksheet.cell(row=target_row, column=columns["Status"]).value = "Active"
        worksheet.cell(
            row=target_row, column=columns["Registered_Date"]
        ).value = registered_date

        try:
            workbook.save(EXCEL_PATH)
        except Exception as exc:
            raise RegistrationError(f"Failed to save attendance workbook: {exc}") from exc

    finally:
        if workbook is not None:
            workbook.close()

    return {
        "Employee_ID": new_employee_id,
        "Employee_Name": employee.get("Employee_Name"),
        "Department": employee.get("Department"),
        "Face_Registered": "No",
        "Status": "Active",
        "Registered_Date": registered_date,
    }